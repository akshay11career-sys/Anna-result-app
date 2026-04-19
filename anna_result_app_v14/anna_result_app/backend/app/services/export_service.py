"""
Export Service
Generates Excel and PDF reports from analytics data.
"""

import io
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from app.core.config import settings


# ──────────────────────────────────────────────
# COLOR CONSTANTS
# ──────────────────────────────────────────────

EXCEL_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
EXCEL_SUBHEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
EXCEL_ALT_FILL = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")
EXCEL_FAIL_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
EXCEL_PASS_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _apply_header(ws, row, cols, fill=EXCEL_HEADER_FILL, font=WHITE_FONT):
    for col, value in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

def export_excel(
    students: List[Dict],
    summary: Dict,
    grade_distribution: Dict,
    arrear_data: Dict,
    yearwise_data: Dict,
    filename: str = None,
) -> str:
    """
    Create a multi-sheet Excel report.
    Returns the file path.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, summary)

    # ── Sheet 2: Student Results ──
    ws_students = wb.create_sheet("Student Results")
    _write_students_sheet(ws_students, students)

    # ── Sheet 3: Grade Distribution ──
    ws_grades = wb.create_sheet("Grade Distribution")
    _write_grade_dist_sheet(ws_grades, grade_distribution)

    # ── Sheet 4: Arrear Analysis ──
    ws_arrears = wb.create_sheet("Arrear Analysis")
    _write_arrear_sheet(ws_arrears, arrear_data)

    # ── Sheet 5: Year-wise Analysis ──
    ws_year = wb.create_sheet("Year-wise Analysis")
    _write_yearwise_sheet(ws_year, yearwise_data)

    # Save
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"anna_result_report_{ts}.xlsx"

    path = os.path.join(settings.EXPORT_DIR, filename)
    wb.save(path)
    return path


def _write_summary_sheet(ws, summary: Dict):
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "ANNA UNIVERSITY RESULT ANALYSIS - SUMMARY"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = EXCEL_HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center")

    row = 3
    items = [
        ("Total Students", summary.get("total_students", 0)),
        ("Current Students", summary.get("current_students", 0)),
        ("Past Arrear Students", summary.get("past_arrear_students", 0)),
        ("Pass Count", summary.get("pass_count", 0)),
        ("Fail Count", summary.get("fail_count", 0)),
        ("Pass %", f"{summary.get('pass_percentage', 0):.2f}%"),
        ("Fail %", f"{summary.get('fail_percentage', 0):.2f}%"),
        ("Average Percentage", f"{summary.get('average_percentage', 0):.2f}%"),
        ("Average CGPA", f"{summary.get('average_cgpa', 0):.2f}"),
        ("Total Arrears", summary.get("total_arrears", 0)),
    ]

    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = HEADER_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    _auto_width(ws)


def _write_students_sheet(ws, students: List[Dict]):
    headers = ["Register No", "Name", "Regulation", "Semester", "Year", "Total Subjects",
               "Arrears", "Percentage", "CGPA", "Type"]
    _apply_header(ws, 1, headers)

    for i, s in enumerate(students, 2):
        fill = EXCEL_FAIL_FILL if s.get("arrear_count", 0) > 0 else None
        row_data = [
            s.get("register_number"),
            s.get("name"),
            s.get("regulation"),
            s.get("semester"),
            s.get("year_label"),
            s.get("total_subjects"),
            s.get("arrear_count"),
            f"{s.get('percentage', 0):.2f}%",
            f"{s.get('cgpa', 0):.2f}",
            "Current" if s.get("is_current_student") else "Past Arrear",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if i % 2 == 0 and not fill:
                cell.fill = EXCEL_ALT_FILL

    _auto_width(ws)


def _write_grade_dist_sheet(ws, grade_distribution: Dict):
    ws.merge_cells("A1:H1")
    ws["A1"].value = "SUBJECT-WISE GRADE DISTRIBUTION"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = EXCEL_HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    subject_wise = grade_distribution.get("subject_wise", [])
    if not subject_wise:
        return

    # Get all unique grades
    all_grades = sorted(set(
        g for s in subject_wise for g in s.get("grade_counts", {}).keys()
    ))

    headers = ["Subject Code", "Subject Name", "Total Students", "Pass", "Fail",
               "Pass %", "Fail %"] + all_grades
    _apply_header(ws, 3, headers)

    for i, subj in enumerate(subject_wise, 4):
        row_data = [
            subj.get("subject_code"),
            subj.get("subject_name", ""),
            subj.get("total_students"),
            subj.get("pass_count"),
            subj.get("fail_count"),
            f"{subj.get('pass_percentage', 0):.1f}%",
            f"{subj.get('fail_percentage', 0):.1f}%",
        ]
        for g in all_grades:
            row_data.append(subj.get("grade_counts", {}).get(g, 0))

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = EXCEL_ALT_FILL

    _auto_width(ws)


def _write_arrear_sheet(ws, arrear_data: Dict):
    headers = ["Register No", "Name", "Regulation", "Semester", "Arrear Count",
               "Failed Subjects", "Type"]
    _apply_header(ws, 1, headers)

    arrear_students = arrear_data.get("arrear_students", [])
    for i, s in enumerate(arrear_students, 2):
        failed = s.get("failed_subjects", [])
        failed_codes = ", ".join(f["subject_code"] for f in failed)
        row_data = [
            s.get("register_number"),
            s.get("name"),
            s.get("regulation"),
            s.get("semester"),
            s.get("arrear_count"),
            failed_codes,
            "Current" if s.get("is_current_student") else "Past",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = THIN_BORDER
            cell.fill = EXCEL_FAIL_FILL if s.get("arrear_count", 0) > 0 else EXCEL_ALT_FILL

    _auto_width(ws)


def _write_yearwise_sheet(ws, yearwise_data: Dict):
    headers = ["Year", "Total Students", "Pass", "Fail", "Pass %",
               "Avg Percentage", "Avg CGPA"]
    _apply_header(ws, 1, headers)

    for i, y in enumerate(yearwise_data.get("years", []), 2):
        row_data = [
            y.get("year_label"),
            y.get("total_students"),
            y.get("pass_count"),
            y.get("fail_count"),
            f"{y.get('pass_percentage', 0):.2f}%",
            f"{y.get('average_percentage', 0):.2f}%",
            f"{y.get('average_cgpa', 0):.2f}",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = EXCEL_ALT_FILL

    _auto_width(ws)


# ──────────────────────────────────────────────
# PDF EXPORT (ReportLab)
# ──────────────────────────────────────────────

def export_pdf_report(
    students: List[Dict],
    summary: Dict,
    grade_distribution: Dict,
    filename: str = None,
) -> str:
    """
    Generate a formatted PDF report.
    Returns the file path.
    """
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"anna_result_report_{ts}.pdf"

    path = os.path.join(settings.EXPORT_DIR, filename)
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#1E3A5F"),
        alignment=TA_CENTER, spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=11, textColor=colors.HexColor("#2E86AB"),
        alignment=TA_CENTER, spaceAfter=6
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor("#1E3A5F"),
        spaceBefore=12, spaceAfter=6
    )

    # Title
    story.append(Paragraph("Anna University Result Analysis Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1E3A5F")))
    story.append(Spacer(1, 0.2*inch))

    # Summary table
    story.append(Paragraph("Summary", section_style))
    summary_data = [
        ["Metric", "Value"],
        ["Total Students", str(summary.get("total_students", 0))],
        ["Pass Count", str(summary.get("pass_count", 0))],
        ["Fail Count", str(summary.get("fail_count", 0))],
        ["Pass %", f"{summary.get('pass_percentage', 0):.2f}%"],
        ["Average Percentage", f"{summary.get('average_percentage', 0):.2f}%"],
        ["Average CGPA", f"{summary.get('average_cgpa', 0):.2f}"],
    ]
    _pdf_table(story, summary_data, col_widths=[3*inch, 2*inch])

    story.append(Spacer(1, 0.2*inch))

    # Student table (first 50)
    story.append(Paragraph("Student Results (Top 50)", section_style))
    student_headers = ["Reg No", "Name", "Regulation", "Sem", "Arrears", "Percentage", "CGPA"]
    student_rows = [student_headers]
    for s in students[:50]:
        student_rows.append([
            s.get("register_number", ""),
            s.get("name", "")[:25],
            s.get("regulation", ""),
            str(s.get("semester", "")),
            str(s.get("arrear_count", 0)),
            f"{s.get('percentage', 0):.1f}%",
            f"{s.get('cgpa', 0):.2f}",
        ])
    _pdf_table(story, student_rows)

    doc.build(story)
    return path


def _pdf_table(story, data: List[List], col_widths=None):
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2FB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT", (0, 0), (-1, -1), 16),
    ]))
    story.append(table)


# ──────────────────────────────────────────────
# CLASS RESULT ANALYSIS — Excel Sheet
# ──────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────────
# CLASS RESULT ANALYSIS — Excel Sheet (Anna University Format)
# Matches official AURCM "Result Analysis" sheet exactly
# ──────────────────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────────
# CLASS RESULT ANALYSIS — Excel Sheet (Anna University Format, Year-wise)
# ──────────────────────────────────────────────────────────────────────────────

def _write_year_block(ws, year_block: dict, start_row: int) -> int:
    """
    Write one year's result block.
    Returns the next available row number after this block.
    """
    yr       = year_block.get("year_label", "")
    sem      = year_block.get("semester",   0)
    strength = year_block.get("class_strength", 0)
    passed   = year_block.get("total_passed",   0)
    failed   = year_block.get("total_failed",   0)
    pct      = year_block.get("overall_pass_percentage", 0)
    subjects = year_block.get("subjects", [])

    r = start_row

    # Year heading row
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value     = f"{yr}  —  Semester {sem}  (Current Students Only)"
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = EXCEL_SUBHEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # Column headers
    headers = [
        "S.No", "Subject Code", "Subject Name",
        "Class Strength", "No. of Students Appeared",
        "No. of Students Passed", "No. of Students Failed", "Pass Percentage",
    ]
    _apply_header(ws, r, headers)
    ws.row_dimensions[r].height = 36
    r += 1

    # Subject data rows
    for subj in subjects:
        p = subj.get("pass_percentage", 0)
        row_data = [
            subj.get("sno", ""),
            subj.get("subject_code", ""),
            subj.get("subject_name", ""),
            subj.get("class_strength", 0),
            subj.get("appeared", 0),
            subj.get("passed",   0),
            subj.get("failed",   0),
            f"{p:.2f}%",
        ]
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if ci != 3 else "left",
                vertical="center"
            )
            if r % 2 == 0:
                cell.fill = EXCEL_ALT_FILL
            # Colour-code pass % (col 8)
            if ci == 8:
                if p >= 75:
                    cell.fill = EXCEL_PASS_FILL
                    cell.font = Font(bold=True, color="006100", size=10)
                elif p >= 50:
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    cell.font = Font(bold=True, color="7D6608", size=10)
                else:
                    cell.fill = EXCEL_FAIL_FILL
                    cell.font = Font(bold=True, color="C00000", size=10)
            # Highlight non-zero failed count (col 7)
            if ci == 7 and value and int(value) > 0:
                cell.fill = EXCEL_FAIL_FILL
                cell.font = Font(bold=True, color="C00000", size=10)
        ws.row_dimensions[r].height = 18
        r += 1

    # Summary rows for this year
    summary_items = [
        ("Total No. of Class Strength",      strength, None),
        ("Total No. of Pass Strength",        passed,  EXCEL_PASS_FILL),
        ("Total No. of Fail Strength",        failed,  EXCEL_FAIL_FILL if failed > 0 else None),
        ("Pass Percentage (Current Sem Only)",f"{pct:.2f}%",
         EXCEL_PASS_FILL if pct >= 75 else EXCEL_FAIL_FILL),
    ]
    for label, value, fill in summary_items:
        ws.merge_cells(f"A{r}:D{r}")
        lc = ws[f"A{r}"]
        lc.value     = label
        lc.font      = Font(bold=True, size=10, color="FFFFFF")
        lc.fill      = EXCEL_SUBHEADER_FILL
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = THIN_BORDER
        vc = ws.cell(row=r, column=5, value=value)
        vc.font      = Font(bold=True, size=12)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = THIN_BORDER
        if fill:
            vc.fill = fill
            if fill == EXCEL_PASS_FILL:
                vc.font = Font(bold=True, size=12, color="006100")
            elif fill == EXCEL_FAIL_FILL:
                vc.font = Font(bold=True, size=12, color="C00000")
        ws.row_dimensions[r].height = 18
        r += 1

    # Blank separator
    r += 1
    return r


def _setup_year_sheet(ws, year_block: dict):
    """
    Write one year's result on its own dedicated sheet.
    Includes institute header, subject table, and AU-style footer
    (Total Class Strength / Total Pass Strength / Pass Percentage).
    """
    from openpyxl.utils import get_column_letter

    yr       = year_block.get("year_label", "")
    sem      = year_block.get("semester", 0)
    strength = year_block.get("class_strength", 0)
    passed   = year_block.get("total_passed", 0)
    failed   = year_block.get("total_failed", 0)
    pct      = year_block.get("overall_pass_percentage", 0.0)
    subjects = year_block.get("subjects", [])

    col_widths = [6, 14, 42, 15, 22, 20, 20, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Institute header ──────────────────────────────────────────────────────
    def _hdr(row, text, size=11, bold=True, italic=False, fill=True, color="FFFFFF"):
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]
        c.value     = text
        c.font      = Font(bold=bold, italic=italic, size=size, color=color if fill else "000000")
        c.fill      = EXCEL_HEADER_FILL if fill else PatternFill(fill_type=None)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    _hdr(1, "ANNA UNIVERSITY – REGIONAL CAMPUS MADURAI", size=13)
    _hdr(2, "Department of Electronics and Communication Engineering")
    _hdr(3, f"B.E. – ECE (Full Time)  ·  {yr}  ·  Semester {sem}  ·  Nov-Dec Examination",
         fill=False, color="000000")
    _hdr(4, "Subject-wise Class Result Analysis  (Current Semester Subjects Only · Odd Sem)",
         fill=False, color="000000", italic=True, size=10)

    ws.row_dimensions[5].height = 6  # spacer

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        "S.No", "Subject Code", "Subject Name",
        "Class Strength", "No. of Students Appeared",
        "No. of Students Passed", "No. of Students Failed", "Pass Percentage",
    ]
    _apply_header(ws, 6, headers)
    ws.row_dimensions[6].height = 36

    # ── Subject rows ──────────────────────────────────────────────────────────
    r = 7
    for sno, subj in enumerate(subjects, 1):
        p = subj.get("pass_percentage", 0)
        row_data = [
            sno,
            subj.get("subject_code", ""),
            subj.get("subject_name", "") or subj.get("subject_code", ""),
            subj.get("class_strength", 0),
            subj.get("appeared", 0),
            subj.get("passed", 0),
            subj.get("failed", 0),
            f"{p:.2f}%",
        ]
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if ci != 3 else "left",
                vertical="center", wrap_text=(ci == 3)
            )
            if r % 2 == 0:
                cell.fill = EXCEL_ALT_FILL
            if ci == 8:
                if p >= 75:
                    cell.fill = EXCEL_PASS_FILL
                    cell.font = Font(bold=True, color="006100", size=10)
                elif p >= 50:
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    cell.font = Font(bold=True, color="7D6608", size=10)
                else:
                    cell.fill = EXCEL_FAIL_FILL
                    cell.font = Font(bold=True, color="C00000", size=10)
            if ci == 7 and value and int(value) > 0:
                cell.fill = EXCEL_FAIL_FILL
                cell.font = Font(bold=True, color="C00000", size=10)
        ws.row_dimensions[r].height = 18
        r += 1

    # ── AU-style footer: Class Strength / Pass Strength / Pass % ─────────────
    data_start = 7
    data_end   = r - 1

    def _footer_row(label, value, val_fill=None, val_font_color="000000"):
        nonlocal r
        # Label spans cols A–C
        ws.merge_cells(f"A{r}:C{r}")
        lc = ws[f"A{r}"]
        lc.value     = label
        lc.font      = Font(bold=True, size=10, color="FFFFFF")
        lc.fill      = EXCEL_SUBHEADER_FILL
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = THIN_BORDER
        # Value in col D
        vc = ws.cell(row=r, column=4, value=value)
        vc.font      = Font(bold=True, size=12, color=val_font_color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = THIN_BORDER
        if val_fill:
            vc.fill = val_fill
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1  # blank gap after data
    _footer_row("Total No. of Class Strength",
                strength)
    _footer_row("Total No. of Pass Strength",
                passed, EXCEL_PASS_FILL, "006100")
    _footer_row("Pass Percentage (Current Semester Only)",
                f"{pct:.2f}%",
                EXCEL_PASS_FILL if pct >= 75 else (
                    PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    if pct >= 50 else EXCEL_FAIL_FILL
                ),
                "006100" if pct >= 75 else ("7D6608" if pct >= 50 else "C00000"))

    r += 1  # spacer
    # legend
    ws.merge_cells(f"A{r}:H{r}")
    lc = ws[f"A{r}"]
    lc.value = "Pass % Guide:  Green ≥ 75%  |  Yellow 50–74%  |  Red < 50%"
    lc.font  = Font(italic=True, size=9, color="595959")


def _write_class_result_sheet(ws, class_result: dict):
    """Summary sheet listing all years together (kept for backward compat)."""
    from openpyxl.utils import get_column_letter
    ws.merge_cells("A1:H1")
    ws["A1"].value     = "ANNA UNIVERSITY – REGIONAL CAMPUS MADURAI"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = EXCEL_HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"].value     = "Department of Electronics and Communication Engineering"
    ws["A2"].font      = Font(bold=True, size=11, color="FFFFFF")
    ws["A2"].fill      = EXCEL_HEADER_FILL
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"].value     = "B.E. – ECE (Full Time) · Year-wise Class Result Analysis · Nov-Dec Exam · Odd Semesters Only"
    ws["A3"].font      = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:H4")
    ws["A4"].value     = "Note: Only current odd-semester subjects are included. Even/arrear semester subjects are excluded."
    ws["A4"].font      = Font(italic=True, size=9, color="595959")
    ws["A4"].alignment = Alignment(horizontal="left")

    col_widths = [6, 14, 40, 15, 22, 20, 20, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    current_row = 6
    for yb in class_result.get("year_blocks", []):
        current_row = _write_year_block(ws, yb, current_row)

    ws.merge_cells(f"A{current_row}:H{current_row}")
    lc = ws[f"A{current_row}"]
    lc.value = "Pass % Guide:  Green ≥ 75%  |  Yellow 50–74%  |  Red < 50%"
    lc.font  = Font(italic=True, size=9, color="595959")


def export_excel_with_class_result(
    students, summary, grade_distribution,
    arrear_data, yearwise_data, class_result_data,
    filename=None,
) -> str:
    """Generate Excel with all sheets including per-year Class Result sheets."""
    wb = Workbook()

    ws1 = wb.active;             ws1.title = "Summary"
    _write_summary_sheet(ws1, summary)

    ws2 = wb.create_sheet("Student Results")
    _write_students_sheet(ws2, students)

    ws3 = wb.create_sheet("Grade Distribution")
    _write_grade_dist_sheet(ws3, grade_distribution)

    ws4 = wb.create_sheet("Arrear Analysis")
    _write_arrear_sheet(ws4, arrear_data)

    ws5 = wb.create_sheet("Year-wise Analysis")
    _write_yearwise_sheet(ws5, yearwise_data)

    # Combined class result sheet (all years together)
    ws6 = wb.create_sheet("Class Result Analysis")
    _write_class_result_sheet(ws6, class_result_data)

    # ── Individual year sheets (one per year) ─────────────────────────────────
    year_short = {"1st Year": "I Year", "2nd Year": "II Year",
                  "3rd Year": "III Year", "4th Year": "IV Year"}
    for yb in class_result_data.get("year_blocks", []):
        label     = yb.get("year_label", "Year")
        sheet_name = year_short.get(label, label)  # e.g. "I Year"
        ws_yr = wb.create_sheet(sheet_name)
        _setup_year_sheet(ws_yr, yb)

    if not filename:
        from datetime import datetime
        filename = f"anna_result_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    import os
    from app.core.config import settings
    path = os.path.join(settings.EXPORT_DIR, filename)
    wb.save(path)
    return path
